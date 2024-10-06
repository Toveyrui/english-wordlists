import pandas as pd
import requests
import hashlib
import time
from openpyxl import Workbook, load_workbook
import eng_to_ipa as ipa  # Import the eng_to_ipa library

# 读取Excel文件，没有标题行
input_file = r'C:\Users\reiky\OneDrive\ドキュメント\课本\6级+托福+考研\your_file.xlsx'
df = pd.read_excel(input_file, header=None)

# 确定包含单词的列（第0列）
words_column = df.columns[0]

# 有道API信息
app_id = '4a48a28f6024bb59'  # 应用ID
app_key = 'yVqLpmIIkicQXFZ5l85QDff66N9KWfB3'  # 应用密钥

def generate_sign(word, salt):
    """
    生成请求签名
    """
    sign_str = app_id + word + str(salt) + app_key
    return hashlib.md5(sign_str.encode('utf-8')).hexdigest()

def get_word_info(word):
    """
    获取单词的发音、翻译和词性信息
    """
    url = 'https://openapi.youdao.com/api'
    salt = 12345  # 固定随机数
    sign = generate_sign(word, salt)  # 生成签名
    params = {
        'q': word,
        'from': 'en',
        'to': 'zh-CHS',
        'appKey': app_id,
        'salt': salt,
        'sign': sign
    }
    for attempt in range(6):  # 最多尝试6次
        try:
            response = requests.get(url, params=params)  # 发送请求
            if response.status_code == 200:  # 检查响应状态码
                data = response.json()  # 解析响应为JSON
                print(f"API response for {word}: {data}")
                if data.get('errorCode') == '0':  # 检查错误码
                    # 提取音标（如果存在）
                    phonetic = data.get('basic', {}).get('phonetic', '')
                    
                    # 如果 'basic' 不存在，尝试从其他字段提取音标
                    if not phonetic:
                        phonetic = data.get('speakUrl', '')

                    # 提取翻译
                    translation = ', '.join(data.get('translation', []))
                    
                    # 提取词性
                    explains = data.get('basic', {}).get('explains', [])
                    part_of_speech = ', '.join(explains)
                    
                    # 获取IPA音标
                    ipa_transcription = ipa.convert(word)
                    
                    return ipa_transcription, translation, part_of_speech
                elif data.get('errorCode') == '411':
                    print("Rate limit reached, waiting for 3 seconds...")
                    time.sleep(1)  # 等待3秒
            else:
                print(f"Error fetching data for {word}, status code: {response.status_code}")
        except Exception as e:
            print(f"Exception for {word}: {e}")
        time.sleep(0.5)  # 等待0.5秒后重试
    return '', '', ''  # 返回空字符串

output_file = 'output_with_translations.xlsx'  # Ensure this is defined before use

try:
    workbook = load_workbook(output_file)  # 尝试加载现有文件
    sheet = workbook.active
    start_row = workbook.active.max_row + 1
except FileNotFoundError:
    workbook = Workbook()  # 创建新文件
    sheet = workbook.active
    # 添加标题行
    sheet.cell(row=1, column=1, value='Word')
    sheet.cell(row=1, column=2, value='Phonetic')
    sheet.cell(row=1, column=3, value='Translation')
    sheet.cell(row=1, column=4, value='Part of Speech')
    start_row = 2
    workbook.save(output_file)  # 保存文件

for idx, word in enumerate(df[words_column], start=start_row):
    print(f"Processing word {idx - start_row + 1}/{len(df)}: {word}")
    phonetic, translation, part_of_speech = get_word_info(word)  # 获取单词信息
    
    # 将原单词、发音、翻译和词性结果保存到Excel
    sheet.cell(row=idx, column=1, value=word)  # 写入单词
    sheet.cell(row=idx, column=2, value=phonetic)  # 写入发音
    sheet.cell(row=idx, column=3, value=translation)  # 写入翻译
    sheet.cell(row=idx, column=4, value=part_of_speech)  # 写入词性
    
    workbook.save(output_file)  # 保存文件

print("处理完成，结果已保存到 output_with_translations.xlsx")

